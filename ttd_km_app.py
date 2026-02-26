"""
TTD Kaplan-Meier Analysis Tool
==============================
A Streamlit-based GUI for automated Time to Discontinuation (TTD)
Kaplan-Meier analysis of oncology treatment lines.

Author: Generated for Prashanth
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
import datetime
from lifelines import KaplanMeierFitter
from lifelines.statistics import logrank_test
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ──────────────────────────────────────────────
# Page config
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="TTD Kaplan-Meier Analyzer",
    page_icon="📊",
    layout="wide",
)

# ──────────────────────────────────────────────
# Custom CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 2rem;
        font-weight: 700;
        color: #1a1a2e;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1rem;
        color: #555;
        margin-bottom: 1.5rem;
    }
    .metric-card {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 1.2rem;
        text-align: center;
        border: 1px solid #e9ecef;
    }
    .metric-value {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1a1a2e;
    }
    .metric-label {
        font-size: 0.85rem;
        color: #777;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 20px;
    }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════
# CORE FUNCTIONS
# ══════════════════════════════════════════════

def load_data(uploaded_file) -> pd.DataFrame:
    """Load CSV or Excel file into a DataFrame."""
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    elif name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file)
    else:
        raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")

    # Normalize column names: strip whitespace, lowercase
    df.columns = df.columns.str.strip().str.lower()

    # ── Ensure required columns exist ──
    required_core = ["mpi_id", "lot", "regimen", "start_date", "end_date"]
    missing = [c for c in required_core if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    # ── Parse date columns ──
    for col in ["start_date", "end_date", "last_visit_date", "death_date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    # ── Synthesize last_visit_date if absent ──
    if "last_visit_date" not in df.columns:
        np.random.seed(42)
        max_end = df["end_date"].dropna().max()
        random_offsets = pd.to_timedelta(np.random.randint(30, 365, size=len(df)), unit="D")
        df["last_visit_date"] = df["end_date"] + random_offsets
        # Cap at a reasonable date
        if pd.notna(max_end):
            cap = max_end + pd.Timedelta(days=730)
            df["last_visit_date"] = df["last_visit_date"].clip(upper=cap)
        st.info("ℹ️ `last_visit_date` column was not found — synthetic dates have been generated.")

    # ── Synthesize death_date if absent ──
    if "death_date" not in df.columns:
        np.random.seed(99)
        # ~20% of patients get a death date
        mask = np.random.rand(len(df)) < 0.20
        random_offsets = pd.to_timedelta(np.random.randint(60, 600, size=len(df)), unit="D")
        df["death_date"] = pd.NaT
        df.loc[mask, "death_date"] = df.loc[mask, "end_date"] + random_offsets[mask]
        st.info("ℹ️ `death_date` column was not found — synthetic dates have been generated for ~20% of patients.")

    # ── Compute duration columns if missing ──
    if "duration_days" not in df.columns:
        df["duration_days"] = (df["end_date"] - df["start_date"]).dt.days
    if "duration_months" not in df.columns:
        df["duration_months"] = df["duration_days"] / 30.44

    return df


def extract_unique_drugs(df: pd.DataFrame) -> list:
    """Extract sorted unique drug names from the regimen column."""
    drugs = set()
    for val in df["regimen"].dropna().unique():
        for drug in str(val).split(","):
            cleaned = drug.strip()
            if cleaned:
                drugs.add(cleaned)
    return sorted(drugs, key=str.lower)


def filter_data(
    df: pd.DataFrame,
    study_start: datetime.date,
    study_end: datetime.date,
    selected_drugs: list,
    selected_lot: str,
) -> pd.DataFrame:
    """Filter the dataset based on study period, regimen, and LOT."""
    filtered = df.copy()

    # Study period filter on start_date
    filtered = filtered[
        (filtered["start_date"] >= pd.Timestamp(study_start))
        & (filtered["start_date"] <= pd.Timestamp(study_end))
    ]

    # Regimen filter (case-insensitive partial match)
    if selected_drugs:
        def matches_any_drug(regimen_str):
            if pd.isna(regimen_str):
                return False
            regimen_lower = str(regimen_str).lower()
            return any(drug.lower() in regimen_lower for drug in selected_drugs)

        filtered = filtered[filtered["regimen"].apply(matches_any_drug)]

    # LOT filter
    if selected_lot and selected_lot != "All":
        filtered = filtered[filtered["lot"].astype(str).str.strip().str.upper() == selected_lot.upper()]

    return filtered.reset_index(drop=True)


def derive_events(
    df: pd.DataFrame,
    threshold_days: int,
    view: str,
    study_start: datetime.date,
    study_end: datetime.date,
) -> pd.DataFrame:
    """
    Determine event (1 = discontinuation) vs censor (0) for each record.
    Returns the dataframe with added columns:
        has_next_lot, event_flag, event_or_censor_date, time_to_event_days
    """
    result = df.copy()

    # ── Identify if patient has a next LOT ──
    # Sort by patient and lot, then flag rows where a higher lot exists
    result["lot_num"] = (
        result["lot"]
        .astype(str)
        .str.extract(r"(\d+)", expand=False)
        .astype(float)
    )

    # For each record, check if there is a record with the same mpi_id and higher lot
    # Use combined_div_mpi_id if available, else mpi_id
    id_col = "combined_div_mpi_id" if "combined_div_mpi_id" in result.columns else "mpi_id"

    patient_max_lot = result.groupby(id_col)["lot_num"].transform("max")
    result["has_next_lot"] = result["lot_num"] < patient_max_lot

    # ── Gap from end_date to last_visit_date ──
    result["gap_days"] = (result["last_visit_date"] - result["end_date"]).dt.days

    # ── Derive event flag ──
    event_flags = []
    event_dates = []

    for _, row in result.iterrows():
        # Condition 1: has next LOT
        if row["has_next_lot"]:
            event_flags.append(1)
            event_dates.append(row["end_date"])
            continue

        # Condition 2: gap exceeds threshold
        if pd.notna(row["gap_days"]) and row["gap_days"] > threshold_days:
            event_flags.append(1)
            event_dates.append(row["end_date"])
            continue

        # Condition 3 (Death View only): death within study period
        if view == "Death View":
            death = row.get("death_date")
            if pd.notna(death):
                if pd.Timestamp(study_start) <= death <= pd.Timestamp(study_end):
                    event_flags.append(1)
                    event_dates.append(death)
                    continue

        # Otherwise censored
        event_flags.append(0)
        event_dates.append(row["last_visit_date"])

    result["event_flag"] = event_flags
    result["event_or_censor_date"] = pd.to_datetime(event_dates)

    # ── Time to event ──
    result["time_to_event_days"] = (
        result["event_or_censor_date"] - result["start_date"]
    ).dt.days

    # Remove rows with invalid durations
    result = result[result["time_to_event_days"] > 0].reset_index(drop=True)

    return result


def generate_km_dataset(df: pd.DataFrame) -> pd.DataFrame:
    """Create the clean KM-ready dataset."""
    id_col = "combined_div_mpi_id" if "combined_div_mpi_id" in df.columns else "mpi_id"
    cols = [
        id_col, "lot", "regimen", "start_date", "end_date",
        "last_visit_date", "death_date",
        "event_flag", "event_or_censor_date", "time_to_event_days",
    ]
    # Add optional columns if present
    for c in ["division_mask", "metastatic", "duration_days", "duration_months"]:
        if c in df.columns:
            cols.append(c)
    available = [c for c in cols if c in df.columns]
    km_df = df[available].copy()
    km_df["time_to_event_months"] = km_df["time_to_event_days"] / 30.44
    return km_df


def compute_summary_table(km_dataset: pd.DataFrame, regimen_label: str = "All") -> pd.DataFrame:
    """Compute summary statistics and survival probabilities at fixed months."""
    T = km_dataset["time_to_event_days"].values
    E = km_dataset["event_flag"].values

    kmf = KaplanMeierFitter()
    kmf.fit(T, event_observed=E)

    n_total = len(km_dataset)
    n_events = int(E.sum())
    n_censored = n_total - n_events

    # Median TTD
    median_ttd = kmf.median_survival_time_
    if np.isinf(median_ttd):
        median_str = "Not reached"
        median_months = "NR"
    else:
        median_str = f"{median_ttd:.0f} days"
        median_months = f"{median_ttd / 30.44:.1f}"

    # Survival probabilities at fixed months
    month_points = [3, 6, 9, 12, 15, 18, 21, 24]
    surv_probs = {}
    ci_lower = {}
    ci_upper = {}

    for m in month_points:
        t_days = m * 30.44
        # Get survival probability at this time
        idx = kmf.survival_function_.index
        valid = idx[idx <= t_days]
        if len(valid) > 0:
            prob = kmf.survival_function_.loc[valid.max()].values[0]
            # Confidence intervals
            ci = kmf.confidence_interval_survival_function_
            ci_low = ci.iloc[:, 0].loc[valid.max()]
            ci_up = ci.iloc[:, 1].loc[valid.max()]
            surv_probs[m] = round(prob, 4)
            ci_lower[m] = round(ci_low, 4)
            ci_upper[m] = round(ci_up, 4)
        else:
            surv_probs[m] = 1.0
            ci_lower[m] = 1.0
            ci_upper[m] = 1.0

    rows = []
    for m in month_points:
        rows.append({
            "Regimen": regimen_label,
            "Month": m,
            "Survival Probability": surv_probs[m],
            "95% CI Lower": ci_lower[m],
            "95% CI Upper": ci_upper[m],
        })

    summary_df = pd.DataFrame(rows)

    # Add a header row with overall stats
    meta = pd.DataFrame([{
        "Regimen": regimen_label,
        "N (Total)": n_total,
        "Events": n_events,
        "Censored": n_censored,
        "Median TTD (days)": median_str,
        "Median TTD (months)": median_months,
    }])

    return meta, summary_df


def plot_km_curve(km_dataset: pd.DataFrame, selected_drugs: list) -> plt.Figure:
    """Generate Kaplan-Meier curve(s). Separate curves per regimen group if multiple drugs."""
    fig, ax = plt.subplots(figsize=(10, 6.5))

    colors = plt.cm.tab10.colors

    # If multiple drugs selected, try to plot per regimen group
    if selected_drugs and len(selected_drugs) > 1:
        # Group by which selected drug matches the regimen
        groups = {}
        for _, row in km_dataset.iterrows():
            regimen_str = str(row.get("regimen", "")).lower()
            matched = []
            for drug in selected_drugs:
                if drug.lower() in regimen_str:
                    matched.append(drug)
            label = ", ".join(matched) if matched else "Other"
            groups.setdefault(label, []).append(row)

        kmf_list = []
        for i, (label, rows) in enumerate(sorted(groups.items())):
            grp = pd.DataFrame(rows)
            if len(grp) < 2:
                continue
            kmf = KaplanMeierFitter()
            kmf.fit(
                grp["time_to_event_days"] / 30.44,  # Convert to months for x-axis
                event_observed=grp["event_flag"],
                label=f"{label} (n={len(grp)})",
            )
            kmf.plot_survival_function(
                ax=ax,
                ci_show=True,
                color=colors[i % len(colors)],
                linewidth=2,
            )
            kmf_list.append((label, grp, kmf))

        # Risk table below
        if kmf_list:
            _add_risk_table(ax, kmf_list, fig)
    else:
        # Single curve
        label = selected_drugs[0] if selected_drugs else "All Regimens"
        kmf = KaplanMeierFitter()
        kmf.fit(
            km_dataset["time_to_event_days"] / 30.44,
            event_observed=km_dataset["event_flag"],
            label=f"{label} (n={len(km_dataset)})",
        )
        kmf.plot_survival_function(
            ax=ax,
            ci_show=True,
            color=colors[0],
            linewidth=2,
        )
        _add_risk_table(ax, [(label, km_dataset, kmf)], fig)

    ax.set_xlabel("Time (Months)", fontsize=12, fontweight="bold")
    ax.set_ylabel("Probability of Remaining on Treatment", fontsize=12, fontweight="bold")
    ax.set_title("Kaplan-Meier: Time to Discontinuation (TTD)", fontsize=14, fontweight="bold")
    ax.set_ylim(0, 1.05)
    ax.set_xlim(left=0)
    ax.legend(loc="upper right", fontsize=9, framealpha=0.9)
    ax.grid(True, alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    fig.tight_layout()
    return fig


def _add_risk_table(ax, kmf_list, fig):
    """Add a number-at-risk table below the KM plot."""
    time_points = [0, 3, 6, 9, 12, 15, 18, 21, 24]
    risk_data = []
    for label, grp, kmf in kmf_list:
        at_risk = []
        for t in time_points:
            t_days = t * 30.44
            n = int(((grp["time_to_event_days"]) >= t_days).sum())
            at_risk.append(n)
        risk_data.append((label, at_risk))

    # Adjust layout to make room for the risk table
    fig.subplots_adjust(bottom=0.18 + 0.04 * len(risk_data))

    for i, (label, counts) in enumerate(risk_data):
        for j, (t, n) in enumerate(zip(time_points, counts)):
            ax.text(
                t, -0.12 - i * 0.06, str(n),
                ha="center", va="top", fontsize=8,
                transform=ax.get_xaxis_transform(),
                color=plt.cm.tab10.colors[i % 10],
            )
        ax.text(
            -0.02, -0.12 - i * 0.06,
            label[:20],
            ha="right", va="top", fontsize=8,
            transform=ax.get_yaxis_transform(),
            fontweight="bold",
            color=plt.cm.tab10.colors[i % 10],
        )

    ax.text(
        -0.02, -0.06, "At risk:",
        ha="right", va="top", fontsize=8, fontweight="bold",
        transform=ax.get_yaxis_transform(),
    )


def get_km_curve_data(km_dataset: pd.DataFrame) -> pd.DataFrame:
    """Return the KM survival function data for Excel export."""
    T = km_dataset["time_to_event_days"].values
    E = km_dataset["event_flag"].values

    kmf = KaplanMeierFitter()
    kmf.fit(T, event_observed=E)

    sf = kmf.survival_function_.copy()
    ci = kmf.confidence_interval_survival_function_.copy()

    curve_df = pd.DataFrame({
        "Time (days)": sf.index,
        "Time (months)": sf.index / 30.44,
        "Survival Probability": sf.values.flatten(),
        "95% CI Lower": ci.iloc[:, 0].values,
        "95% CI Upper": ci.iloc[:, 1].values,
    })
    return curve_df


def export_to_excel(km_dataset, curve_data, meta_df, summary_df) -> bytes:
    """Export all outputs to an Excel file with three sheets."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        km_dataset.to_excel(writer, sheet_name="KM_Dataset", index=False)
        curve_data.to_excel(writer, sheet_name="KM_Curve_Data", index=False)

        # Summary sheet: meta info at top, then survival probabilities
        meta_df.to_excel(writer, sheet_name="Summary_Table", index=False, startrow=0)
        summary_df.to_excel(
            writer, sheet_name="Summary_Table", index=False,
            startrow=len(meta_df) + 2,
        )

        # ── Format sheets ──
        wb = writer.book
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            # Auto-width
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    return output.getvalue()


# ══════════════════════════════════════════════
# STREAMLIT UI
# ══════════════════════════════════════════════

st.markdown('<div class="main-header">📊 TTD Kaplan-Meier Analyzer</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="sub-header">Automated Time to Discontinuation analysis for oncology treatment lines</div>',
    unsafe_allow_html=True,
)

# ── Sidebar: Input Controls ──
with st.sidebar:
    st.header("⚙️ Analysis Parameters")

    # 1. File upload
    st.subheader("1. Data Upload")
    uploaded_file = st.file_uploader(
        "Upload LOT data (CSV or Excel)",
        type=["csv", "xlsx", "xls"],
        help="File must contain: mpi_id, lot, regimen, start_date, end_date",
    )

    st.divider()

    # 2. Study period
    st.subheader("2. Study Period")
    col1, col2 = st.columns(2)
    with col1:
        study_start = st.date_input(
            "Start Date",
            value=datetime.date(2020, 1, 1),
            min_value=datetime.date(2000, 1, 1),
        )
    with col2:
        study_end = st.date_input(
            "End Date",
            value=datetime.date(2025, 12, 31),
            min_value=datetime.date(2000, 1, 1),
        )

    st.divider()

    # Placeholders — populated after file load
    st.subheader("3. Regimen Filter")
    drug_placeholder = st.empty()

    st.subheader("4. LOT Filter")
    lot_placeholder = st.empty()

    st.divider()

    # 5. Route selection
    st.subheader("5. Route Selection")
    route = st.radio(
        "Administration route",
        options=["IV/SQ", "Oral"],
        horizontal=True,
        help="Sets the discontinuation gap threshold",
    )
    threshold = 120 if route == "IV/SQ" else 180
    st.caption(f"Gap threshold: **{threshold} days**")

    st.divider()

    # 6. View selection
    st.subheader("6. View Selection")
    view = st.radio(
        "Analysis view",
        options=["Geo View", "Death View"],
        horizontal=True,
        help="Death View additionally considers death date as an event",
    )

    st.divider()

    # 7. Run button
    run_button = st.button("🚀 Run Analysis", use_container_width=True, type="primary")

# ── Main area ──

if uploaded_file is not None:
    try:
        raw_df = load_data(uploaded_file)
        st.success(f"✅ Loaded **{len(raw_df):,}** records with **{raw_df['mpi_id'].nunique():,}** unique patients.")

        # Populate sidebar filters
        all_drugs = extract_unique_drugs(raw_df)
        selected_drugs = drug_placeholder.multiselect(
            "Select drug(s)",
            options=all_drugs,
            default=[],
            help="Case-insensitive partial match. Leave empty for all.",
        )

        lot_values = sorted(raw_df["lot"].dropna().astype(str).unique(), key=lambda x: x)
        selected_lot = lot_placeholder.selectbox(
            "Select LOT",
            options=["All"] + lot_values,
        )

        # Show preview
        with st.expander("📋 Data Preview (first 100 rows)", expanded=False):
            st.dataframe(raw_df.head(100), use_container_width=True, height=300)

    except Exception as e:
        st.error(f"❌ Error loading file: {e}")
        raw_df = None
else:
    raw_df = None
    drug_placeholder.info("Upload a file first")
    lot_placeholder.info("Upload a file first")

    # Show instructions
    st.info(
        "👈 **Upload your LOT data file** in the sidebar to get started.\n\n"
        "The file should contain columns like `mpi_id`, `lot`, `regimen`, "
        "`start_date`, `end_date`, etc."
    )

# ── Run Analysis ──
if run_button and raw_df is not None:
    with st.spinner("Running TTD analysis..."):
        # Step 1: Filter
        filtered = filter_data(raw_df, study_start, study_end, selected_drugs, selected_lot)

        if len(filtered) == 0:
            st.error("⚠️ No records match the selected filters. Please adjust your criteria.")
        else:
            # Step 2: Derive events
            result_df = derive_events(filtered, threshold, view, study_start, study_end)

            if len(result_df) == 0:
                st.error("⚠️ No valid records after event derivation.")
            else:
                # Step 3: Generate KM dataset
                km_dataset = generate_km_dataset(result_df)

                # Step 4: Summary
                regimen_label = ", ".join(selected_drugs) if selected_drugs else "All Regimens"
                meta_df, summary_df = compute_summary_table(km_dataset, regimen_label)

                # Step 5: Plot
                fig = plot_km_curve(km_dataset, selected_drugs)

                # Step 6: KM curve data
                curve_data = get_km_curve_data(km_dataset)

                # ── Display Results ──
                st.markdown("---")
                st.markdown("## 📈 Analysis Results")

                # Metric cards
                n_total = len(km_dataset)
                n_events = int(km_dataset["event_flag"].sum())
                n_censored = n_total - n_events
                median_val = meta_df["Median TTD (days)"].values[0]

                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Total Patients (N)", f"{n_total:,}")
                c2.metric("Events", f"{n_events:,}")
                c3.metric("Censored", f"{n_censored:,}")
                c4.metric("Event Rate", f"{n_events/n_total*100:.1f}%")
                c5.metric("Median TTD", median_val)

                # Tabs
                tab1, tab2, tab3, tab4 = st.tabs([
                    "📉 KM Curve", "📋 KM Dataset", "📊 Summary Table", "💾 Export"
                ])

                with tab1:
                    st.pyplot(fig, use_container_width=True)

                with tab2:
                    st.dataframe(km_dataset, use_container_width=True, height=400)
                    st.caption(f"Showing {len(km_dataset):,} records")

                with tab3:
                    st.markdown("### Overall Statistics")
                    st.dataframe(meta_df, use_container_width=True)
                    st.markdown("### Survival Probabilities at Fixed Months")
                    st.dataframe(summary_df, use_container_width=True)

                with tab4:
                    excel_bytes = export_to_excel(km_dataset, curve_data, meta_df, summary_df)
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=excel_bytes,
                        file_name=f"TTD_KM_Analysis_{datetime.date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
                        use_container_width=True,
                        type="primary",
                    )
                    st.success("Excel file contains 3 sheets: KM_Dataset, KM_Curve_Data, Summary_Table")

elif run_button and raw_df is None:
    st.warning("⚠️ Please upload a data file first.")
